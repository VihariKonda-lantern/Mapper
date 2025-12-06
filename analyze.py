"""
pyright: reportMissingTypeStubs=false, reportUnknownVariableType=false, reportUnknownArgumentType=false, reportUnknownMemberType=false
"""
import zipfile, statistics, re
from pathlib import Path
try:
    from openpyxl import load_workbook  # type: ignore[import-not-found]
except Exception:
    # Fallback for type-checkers when openpyxl isn't in the analysis environment
    load_workbook = None  # type: ignore[assignment]
from typing import List, Tuple, Dict, Any, Optional

# __file__ points to app/analyze.py
BASE_DIR = Path(__file__).resolve().parent.parent  # goes up to claims_mapper_app/
FILE_PATH = BASE_DIR / "data" / "AMN_Lantern_Claims_Extract_2025_01.xlsx"


def human(n: float) -> str:
    for unit in ["B","KB","MB","GB"]:
        if n < 1024.0:
            return f"{n:.1f} {unit}"
        n /= 1024.0
    return f"{n:.1f} TB"

def probe_media(xlsx_path: Path) -> Tuple[List[Tuple[str, int]], int]:
    media: List[Tuple[str, int]] = []
    total: int = 0
    with zipfile.ZipFile(xlsx_path) as z:
        for name in z.namelist():
            if name.startswith("xl/media/"):
                size = len(z.read(name))
                media.append((name, size))
                total += size
    media.sort(key=lambda x: x[1], reverse=True)
    return media, total

def looks_header(vals: List[Any], k: int) -> bool:
    vals = [str(x or "").strip() for x in vals]
    nz = [v for v in vals if v]
    if len(nz) < max(3, int(0.5*k)): 
        return False
    nums = sum(1 for v in nz if re.fullmatch(r"[-+]?\d+(\.\d+)?", v))
    short = sum(1 for v in nz if len(v) <= 30)
    return nums <= max(1, int(0.2*len(nz))) and (short/len(nz)) >= 0.7

def scan_xlsx(xlsx_path: Path, sheet_name: Optional[str] = None, sample_rows: int = 300) -> Dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not available in this environment.")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)  # type: ignore[call-arg]
    ws: Any = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    max_r = ws.max_row  # type: ignore[attr-defined]
    start_r = max(1, max_r - sample_rows + 1)
    widths: List[Tuple[int, int]] = []
    rows_cache: Dict[int, List[Any]] = {}
    max_cell_len: int = 0
    suspicious_cells: List[Tuple[int, int]] = []

    for r in range(start_r, max_r + 1):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))  # type: ignore[attr-defined]
        rows_cache[r] = list(row)
        width = sum(1 for v in row if (v is not None and str(v).strip() != ""))
        widths.append((r, width))
        for v in row:
            if v is None: continue
            s = str(v)
            max_cell_len = max(max_cell_len, len(s))
            if len(s) > 100_000 or "\x00" in s:
                suspicious_cells.append((r, len(s)))

    nz = [w for _, w in widths if w > 0]
    width_mode: Optional[int] = None
    header_row: Optional[int] = None
    if nz:
        try:
            width_mode = statistics.mode(nz)
        except statistics.StatisticsError:
            freq: Dict[int, int] = {}
            for w in nz:
                freq[w] = freq.get(w, 0) + 1
            width_mode = max(freq.items(), key=lambda kv:(kv[1], kv[0]))[0]
        for r, w in widths:
            if w >= max(2, int(0.7*width_mode)) and looks_header(rows_cache[r], width_mode):
                header_row = r
                break
        if not header_row:
            for r, w in widths:
                if w >= max(2, int(0.7*width_mode)):
                    header_row = r
                    break

    report: Dict[str, Any] = {
        "sheet": ws.title,  # type: ignore[attr-defined]
        "total_rows": max_r,
        "width_mode": width_mode,
        "header_row_guess": header_row,
        "max_cell_len": max_cell_len,
        "suspicious_cells": suspicious_cells[:5],
    }
    return report

def main() -> None:
    if not FILE_PATH.exists():
        print("File not found:", FILE_PATH)
        return
    print("Inspecting:", FILE_PATH, f"({human(FILE_PATH.stat().st_size)})")

    try:
        media, total = probe_media(FILE_PATH)
        if media:
            print(f"Embedded media total: {human(total)}")
            for name, sz in media[:5]:
                print("  ", name, "→", human(sz))
        else:
            print("No embedded media found.")
    except Exception as e:
        print("Media probe failed:", e)

    try:
        rpt = scan_xlsx(FILE_PATH)
        print("Scan report:", rpt)
    except Exception as e:
        print("openpyxl scan failed:", e)

if __name__ == "__main__":
    main()
